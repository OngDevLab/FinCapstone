import matplotlib.pyplot as plt
import openpyxl
import numpy as np

wb = openpyxl.load_workbook('Darden Financials.xlsx', data_only=True)

# 1. Revenue Trend 2019-2025
ws = wb['Income Statement ']
years = [2019, 2020, 2021, 2022, 2023, 2024, 2025]
revenues = [8080.1, 7806.9, 7196.1, 9630, 10487.8, 11390, 12076.7]

fig, ax = plt.subplots(figsize=(12, 6))
ax.plot(years, revenues, marker='o', linewidth=3, markersize=10, color='#1f77b4')
ax.set_xlabel('Fiscal Year', fontsize=12)
ax.set_ylabel('Revenue ($M)', fontsize=12)
ax.set_title('Darden Revenue Trend (FY2019-FY2025)', fontsize=14, fontweight='bold')
ax.grid(True, alpha=0.3)
for i, (year, rev) in enumerate(zip(years, revenues)):
    ax.text(year, rev + 200, f'${rev:,.0f}M', ha='center', fontsize=9)
plt.tight_layout()
plt.savefig('revenue_trend_2019_2025.png', dpi=300, bbox_inches='tight')
plt.close()
print("✓ revenue_trend_2019_2025.png")

# 2. Margin Trends Over Time
ws = wb['ratios']
margin_data = {}
for row in ws.iter_rows(min_row=1, values_only=True):
    if row[0] == 'Operating Margin':
        margin_data['Operating'] = [v*100 for v in row[1:8] if v is not None]
    elif row[0] == 'Net Profit Margin':
        margin_data['Net Profit'] = [v*100 for v in row[1:8] if v is not None]
    elif row[0] == 'EBITDA Margin':
        margin_data['EBITDA'] = [v*100 for v in row[1:8] if v is not None]

fig, ax = plt.subplots(figsize=(12, 6))
years_margin = [2019, 2020, 2021, 2022, 2023, 2024, 2025]
ax.plot(years_margin, margin_data['EBITDA'], marker='o', linewidth=2, label='EBITDA Margin')
ax.plot(years_margin, margin_data['Operating'], marker='s', linewidth=2, label='Operating Margin')
ax.plot(years_margin, margin_data['Net Profit'], marker='^', linewidth=2, label='Net Profit Margin')
ax.set_xlabel('Fiscal Year', fontsize=12)
ax.set_ylabel('Margin (%)', fontsize=12)
ax.set_title('Darden Profitability Margins Trend (FY2019-FY2025)', fontsize=14, fontweight='bold')
ax.legend(fontsize=10)
ax.grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig('margin_trends_2019_2025.png', dpi=300, bbox_inches='tight')
plt.close()
print("✓ margin_trends_2019_2025.png")

# 3. Segment Revenue Mix (FY2025)
ws = wb['brand sales']
segments = []
segment_revenue = []
for row in ws.iter_rows(min_row=2, max_row=10, values_only=True):
    if row[0] and row[1]:
        segments.append(row[0].title())
        segment_revenue.append(row[1])

# Group into 4 main segments
segment_groups = {
    'Olive Garden': 5236,
    'LongHorn': 3073.2,
    'Fine Dining': 442.8 + 624.8 + 292.4 + 246.5 + 880,  # Ruth's, Capital, Seasons, Eddie V's, Yard House
    'Other': 742.1 + 475.2 + 173.6  # Cheddar's, Chuy's, Bahama Breeze
}

fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))

# Pie chart
colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728']
wedges, texts, autotexts = ax1.pie(segment_groups.values(), labels=segment_groups.keys(), 
                                     autopct='%1.1f%%', colors=colors, startangle=90)
for autotext in autotexts:
    autotext.set_color('white')
    autotext.set_fontsize(11)
    autotext.set_fontweight('bold')
ax1.set_title('Darden Revenue Mix by Segment (FY2025)', fontsize=12, fontweight='bold')

# Bar chart
ax2.barh(list(segment_groups.keys()), list(segment_groups.values()), color=colors)
ax2.set_xlabel('Revenue ($M)', fontsize=11)
ax2.set_title('Segment Revenue Comparison (FY2025)', fontsize=12, fontweight='bold')
ax2.grid(axis='x', alpha=0.3)
for i, v in enumerate(segment_groups.values()):
    ax2.text(v + 100, i, f'${v:,.0f}M', va='center', fontsize=9)

plt.tight_layout()
plt.savefig('segment_revenue_mix.png', dpi=300, bbox_inches='tight')
plt.close()
print("✓ segment_revenue_mix.png")

# 4. CPI/DPI Trends
ws = wb['Sheet1']
years_macro = []
cpi = []
dpi = []
for row in ws.iter_rows(min_row=1, values_only=True):
    if row[0] is None:
        years_macro = [v for v in row[1:7] if v is not None]
    elif row[0] == 'CPI':
        cpi = [v*100 for v in row[1:7] if v is not None]
    elif row[0] == 'DPI':
        dpi = [v*100 for v in row[1:7] if v is not None]

fig, ax = plt.subplots(figsize=(12, 6))
x = np.arange(len(years_macro))
width = 0.35
ax.bar(x - width/2, cpi, width, label='CPI (Inflation)', color='#d62728')
ax.bar(x + width/2, dpi, width, label='Disposable Personal Income', color='#2ca02c')
ax.set_xlabel('Year', fontsize=12)
ax.set_ylabel('Year-over-Year Change (%)', fontsize=12)
ax.set_title('Consumer Inflation vs. Disposable Income Growth', fontsize=14, fontweight='bold')
ax.set_xticks(x)
ax.set_xticklabels(years_macro)
ax.legend(fontsize=10)
ax.axhline(y=0, color='black', linestyle='-', linewidth=0.8)
ax.grid(axis='y', alpha=0.3)
for i, (c, d) in enumerate(zip(cpi, dpi)):
    ax.text(i - width/2, c + 0.3, f'{c:.1f}%', ha='center', fontsize=8)
    ax.text(i + width/2, d + 0.3, f'{d:.1f}%', ha='center', fontsize=8)
plt.tight_layout()
plt.savefig('cpi_dpi_trends.png', dpi=300, bbox_inches='tight')
plt.close()
print("✓ cpi_dpi_trends.png")

# 5. TTM vs FY2025 Comparison
metrics = ['Revenue', 'Operating Income', 'Net Earnings']
fy2025 = [12076.7, 1362.3, 1049.6]
ttm = [12187.7, 1382.3, 1064.6]

fig, ax = plt.subplots(figsize=(10, 6))
x = np.arange(len(metrics))
width = 0.35
ax.bar(x - width/2, fy2025, width, label='FY2025', color='#1f77b4')
ax.bar(x + width/2, ttm, width, label='TTM (Most Recent 12 Months)', color='#ff7f0e')
ax.set_ylabel('Amount ($M)', fontsize=12)
ax.set_title('Darden: FY2025 vs. Trailing Twelve Months (TTM)', fontsize=14, fontweight='bold')
ax.set_xticks(x)
ax.set_xticklabels(metrics)
ax.legend(fontsize=10)
ax.grid(axis='y', alpha=0.3)
for i, (f, t) in enumerate(zip(fy2025, ttm)):
    ax.text(i - width/2, f + 50, f'${f:,.0f}M', ha='center', fontsize=9)
    ax.text(i + width/2, t + 50, f'${t:,.0f}M', ha='center', fontsize=9)
    pct_change = ((t/f - 1) * 100)
    ax.text(i, max(f, t) + 200, f'+{pct_change:.1f}%', ha='center', fontsize=9, 
            color='green' if pct_change > 0 else 'red', fontweight='bold')
plt.tight_layout()
plt.savefig('ttm_vs_fy2025.png', dpi=300, bbox_inches='tight')
plt.close()
print("✓ ttm_vs_fy2025.png")

print("\n✅ All additional graphs created successfully!")
